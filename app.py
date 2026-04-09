"""
Aplicación web de carrito de compras mayorista - RM KITS
"""
from flask import Flask, render_template, request, jsonify, redirect, session, flash, url_for, send_file, send_from_directory
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import sqlite3
import urllib.parse
import logging
import os
import json
from contextlib import contextmanager
from functools import wraps
from datetime import datetime
from config import Config
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from io import BytesIO
import zipfile
import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Inicializar Flask
app = Flask(__name__)
app.config.from_object(Config)


# Inicializar carpetas de almacenamiento persistente
def init_persistent_storage():
    """Crea las carpetas de almacenamiento persistente si no existen y verifica la DB"""
    logger.info("="*60)
    logger.info("INICIANDO APLICACIÓN RM KITS")
    logger.info("="*60)
    
    try:
        # Loguear configuración de rutas
        logger.info(f"📂 PERSISTENT_DATA_PATH: {Config.PERSISTENT_DATA_PATH}")
        logger.info(f"💾 DATABASE_PATH: {Config.DATABASE_PATH}")
        logger.info(f"🖼️  UPLOAD_FOLDER: {Config.UPLOAD_FOLDER}")
        
        # Crear carpeta principal de datos persistentes
        os.makedirs(Config.PERSISTENT_DATA_PATH, exist_ok=True)
        logger.info(f"✓ Carpeta persistente verificada: {Config.PERSISTENT_DATA_PATH}")
        
        # Crear carpeta de imágenes
        os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
        logger.info(f"✓ Carpeta de imágenes verificada: {Config.UPLOAD_FOLDER}")
        
        # SOLO EN DESARROLLO: migrar archivos desde ubicaciones antiguas
        if Config.PERSISTENT_DATA_PATH != '/data':
            # Migrar imágenes desde static/img si existe
            static_img = os.path.join('static', 'img')
            if os.path.exists(static_img) and os.path.isdir(static_img):
                for filename in os.listdir(static_img):
                    src = os.path.join(static_img, filename)
                    dst = os.path.join(Config.UPLOAD_FOLDER, filename)
                    if os.path.isfile(src) and not os.path.exists(dst):
                        shutil.copy2(src, dst)
                logger.info("✓ Imágenes migradas desde static/img")
            
            # Migrar productos.db desde la raíz si existe
            if os.path.exists('productos.db') and Config.DATABASE_PATH != 'productos.db':
                if not os.path.exists(Config.DATABASE_PATH):
                    shutil.move('productos.db', Config.DATABASE_PATH)
                    logger.info(f"✓ Base de datos movida a: {Config.DATABASE_PATH}")
        
        # VERIFICACIÓN CRÍTICA: Comprobar existencia de la base de datos
        if os.path.exists(Config.DATABASE_PATH):
            db_size = os.path.getsize(Config.DATABASE_PATH)
            logger.info(f"✅ Base de datos encontrada: {Config.DATABASE_PATH} ({db_size:,} bytes)")
            
            # Verificar que la DB tenga contenido
            if db_size < 1024:  # Menos de 1KB es sospechoso
                logger.warning(f"⚠️  ADVERTENCIA: La base de datos parece estar vacía ({db_size} bytes)")
        else:
            logger.error("="*60)
            logger.error("❌ ERROR CRÍTICO: Base de datos NO encontrada")
            logger.error(f"❌ Ruta esperada: {Config.DATABASE_PATH}")
            logger.error("❌ La aplicación arrancará sin la base cargada")
            logger.error("="*60)
            return
        
        logger.info("="*60)
        logger.info("✅ Inicialización completada correctamente")
        logger.info("="*60)
        
    except FileNotFoundError:
        # Re-lanzar el error de DB no encontrada
        raise
    except Exception as e:
        logger.error(f"❌ Error al inicializar almacenamiento persistente: {e}")
        raise


# Inicializar al cargar la aplicación
init_persistent_storage()


# Context manager para manejo seguro de base de datos
@contextmanager
def get_db_connection():
    """Context manager para conexiones a la base de datos"""
    conn = None
    try:
        conn = sqlite3.connect(Config.DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        yield conn
    except sqlite3.Error as e:
        logger.error(f"Error de base de datos: {e}")
        raise
    finally:
        if conn:
            conn.close()


@app.before_request
def before_request():
    """Forzar HTTPS en producción"""
    if request.headers.get("X-Forwarded-Proto") == "http":
        url = request.url.replace("http://", "https://", 1)
        return redirect(url, code=301)


def get_productos():
    """Obtiene todos los productos activos de la base de datos"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM producto WHERE stock > 0 AND activo = 1 ORDER BY codigo DESC")
            filas = cursor.fetchall()
            productos = [dict(fila) for fila in filas]
            return productos
    except Exception as e:
        logger.error(f"Error al obtener productos: {e}")
        return []


@app.route("/")
def index():
    """Página principal con lista de productos"""
    try:
        productos = get_productos()
        
        # Obtener categorías de la base de datos
        categorias = []
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM categoria ORDER BY nombre")
                categorias = [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error al obtener categorías: {e}")
        
        return render_template(
            "index.html",
            productos=productos,
            categorias=categorias,
            config={
                'pedido_minimo': Config.PEDIDO_MINIMO,
                'whatsapp': Config.WHATSAPP_NUMBER
            }
        )
    except Exception as e:
        logger.error(f"Error en página principal: {e}")
        return render_template("error.html", mensaje="Error al cargar productos"), 500


@app.route("/carrito")
def carrito_view():
    """Página del carrito de compras"""
    try:
        # Obtener productos actualizados de la BD para sincronizar precios, stock e imágenes
        # Usar código como clave porque los IDs cambian al reimportar Excel
        productos_actualizados = {}
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, codigo, precio, stock, minimo, multiplo, imagen FROM producto WHERE activo = 1 ORDER BY codigo DESC")
                for row in cursor.fetchall():
                    productos_actualizados[row['codigo']] = {
                        'id': row['id'],
                        'precio': row['precio'],
                        'stock': row['stock'],
                        'minimo': row['minimo'],
                        'multiplo': row['multiplo'],
                        'imagen': row['imagen']
                    }
        except Exception as db_error:
            logger.warning(f"Error al obtener precios actualizados: {db_error}")
        
        return render_template(
            "carrito.html",
            productos_actualizados=productos_actualizados,
            config={
                'pedido_minimo': Config.PEDIDO_MINIMO,
                'whatsapp': Config.WHATSAPP_NUMBER,
                'local_direccion': Config.LOCAL_DIRECCION,
                'local_horarios': Config.LOCAL_HORARIOS,
                'envio_caba': Config.ENVIO_CABA,
                'envio_gba': Config.ENVIO_GBA
            }
        )
    except Exception as e:
        logger.error(f"Error en página de carrito: {e}")
        return render_template("error.html", mensaje="Error al cargar carrito"), 500


@app.route("/enviar_pedido", methods=["POST"])
def enviar_pedido():
    """
    API endpoint para procesar pedidos
    (Actualmente no se usa, pero se mantiene por compatibilidad)
    """
    try:
        datos = request.json
        if not datos:
            return jsonify({"error": "No se recibieron datos"}), 400
        
        total = datos.get("total", 0)
        items = datos.get("items", [])

        if total < Config.PEDIDO_MINIMO:
            return jsonify({
                "error": f"El pedido debe superar los ${Config.PEDIDO_MINIMO:,.0f}"
            }), 400

        # Construir mensaje
        lines = ["Pedido mayorista RM KITS:"]
        for item in items:
            codigo = item.get('codigo', '')
            titulo = item.get('titulo', '')
            cantidad = item.get('cantidad', 0)
            precio = item.get('precio', 0)
            lines.append(f"{codigo} - {titulo} - Cantidad: {cantidad} - Precio unitario: ${precio}")
        lines.append(f"TOTAL: ${total}")

        mensaje = "\n".join(lines)
        encoded = urllib.parse.quote(mensaje)
        url = f"https://wa.me/{Config.WHATSAPP_NUMBER}?text={encoded}"

        return jsonify({"url": url})
    
    except Exception as e:
        logger.error(f"Error al enviar pedido: {e}")
        return jsonify({"error": "Error al procesar el pedido"}), 500


def enviar_email_confirmacion(pedido_id, cliente_nombre, cliente_email, productos_json, total, metodo_entrega, datos_envio=None):
    """Envía email de confirmación de pedido al cliente"""
    try:
        # Si no hay email configurado o el cliente no proporcionó email, no enviar
        if not cliente_email or not Config.MAIL_USERNAME or Config.MAIL_USERNAME == 'tu_email@gmail.com':
            logger.info("Email no configurado o cliente sin email, saltando envío")
            return
        
        # Parsear productos
        productos = json.loads(productos_json) if isinstance(productos_json, str) else productos_json
        
        # Crear mensaje
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'Confirmación de Pedido #{pedido_id} - RM KITS'
        msg['From'] = Config.MAIL_DEFAULT_SENDER
        msg['To'] = cliente_email
        
        # Construir el cuerpo del email en HTML
        html = f"""
        <html>
          <head>
            <style>
              body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
              .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
              .header {{ background-color: #6a1b9a; color: white; padding: 20px; text-align: center; }}
              .content {{ padding: 20px; background-color: #f9f9f9; }}
              .pedido-info {{ background-color: white; padding: 15px; margin: 15px 0; border-left: 4px solid #6a1b9a; }}
              .producto {{ padding: 10px; border-bottom: 1px solid #eee; }}
              .total {{ font-size: 1.3em; font-weight: bold; color: #6a1b9a; margin-top: 20px; padding: 15px; background-color: #f0e6f6; text-align: right; }}
              .footer {{ text-align: center; padding: 20px; color: #666; font-size: 0.9em; }}
            </style>
          </head>
          <body>
            <div class="container">
              <div class="header">
                <h1>¡Gracias por tu pedido!</h1>
              </div>
              
              <div class="content">
                <p>Hola <strong>{cliente_nombre}</strong>,</p>
                
                <p>Hemos recibido tu pedido correctamente. En breve nos comunicaremos con vos para confirmar los detalles y coordinar la entrega.</p>
                
                <div class="pedido-info">
                  <h3>📋 Número de Pedido: #{pedido_id}</h3>
                  <p><strong>Método de entrega:</strong> {metodo_entrega.replace('retiro', 'Retiro en local').replace('envio', 'Envío a domicilio')}</p>
        """
        
        # Agregar datos de envío si aplica
        if metodo_entrega == 'envio' and datos_envio:
            html += f"""
                  <p><strong>Dirección de envío:</strong><br>
                  {datos_envio.get('direccion', '')}<br>
                  {datos_envio.get('localidad', '')}, {datos_envio.get('provincia', '')}<br>
                CP: {datos_envio.get('cp', '')}<br>
                DNI del destinatario: {datos_envio.get('dni_destinatario', datos_envio.get('cuit_destinatario', ''))}</p>
            """
        
        html += """
                </div>
                
                <h3>🛒 Productos:</h3>
        """
        
        # Listar productos
        for p in productos:
            subtotal = p['precio'] * p['cantidad']
            html += f"""
                <div class="producto">
                  <strong>{p['codigo']} - {p['titulo']}</strong><br>
                  Cantidad: {p['cantidad']} × ${p['precio']:,.0f} = ${subtotal:,.0f}
                </div>
            """
        
        html += f"""
                <div class="total">
                  TOTAL: ${total:,.0f}
                </div>
                
                <p style="margin-top: 20px;">Si tenés alguna consulta, podés comunicarte con nosotros:</p>
                <ul>
                  <li><strong>WhatsApp:</strong> +54 9 11 5857-3906</li>
                  <li><strong>Dirección:</strong> Av. Rivadavia 2768, CABA</li>
                </ul>
              </div>
              
              <div class="footer">
                <p>Este es un email automático, por favor no respondas a este mensaje.</p>
                <p>© {datetime.now().year} RM KITS - Todos los derechos reservados</p>
              </div>
            </div>
          </body>
        </html>
        """
        
        # Adjuntar HTML
        part = MIMEText(html, 'html')
        msg.attach(part)
        
        # Enviar email
        with smtplib.SMTP(Config.MAIL_SERVER, Config.MAIL_PORT) as server:
            server.starttls()
            server.login(Config.MAIL_USERNAME, Config.MAIL_PASSWORD)
            server.send_message(msg)
        
        logger.info(f"Email de confirmación enviado a {cliente_email}")
        
    except Exception as e:
        logger.error(f"Error al enviar email de confirmación: {e}")
        # No lanzar error para no interrumpir el flujo del pedido


@app.errorhandler(404)
def page_not_found(e):
    """Página no encontrada"""
    return render_template("error.html", mensaje="Página no encontrada"), 404


@app.errorhandler(500)
def internal_error(e):
    """Error interno del servidor"""
    logger.error(f"Error 500: {e}")
    return render_template("error.html", mensaje="Error interno del servidor"), 500


# =============================================================================
# RUTA TEMPORAL DE RESTAURACIÓN (ELIMINAR DESPUÉS DE USAR)
# =============================================================================

@app.route("/restaurar", methods=["GET", "POST"])
def restaurar_backup():
    """
    RUTA TEMPORAL: Restaura base de datos e imágenes desde un archivo ZIP.
    Extrae productos.db -> /data/productos.db
    Extrae img/* -> /data/img/
    ⚠️ ELIMINAR ESTA RUTA DESPUÉS DE USARLA
    """
    if request.method == "GET":
        # Mostrar formulario de carga
        html = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Restaurar Backup - RM KITS</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    padding: 20px;
                }
                .container {
                    background: white;
                    border-radius: 16px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    padding: 40px;
                    max-width: 600px;
                    width: 100%;
                }
                h1 {
                    color: #333;
                    margin-bottom: 10px;
                    font-size: 28px;
                }
                .warning {
                    background: #fff3cd;
                    border-left: 4px solid #ffc107;
                    color: #856404;
                    padding: 15px;
                    margin: 20px 0;
                    border-radius: 4px;
                    font-size: 14px;
                }
                .info {
                    background: #d1ecf1;
                    border-left: 4px solid #17a2b8;
                    color: #0c5460;
                    padding: 15px;
                    margin: 20px 0;
                    border-radius: 4px;
                    font-size: 14px;
                }
                .info ul {
                    margin: 10px 0 0 20px;
                }
                .form-group {
                    margin: 25px 0;
                }
                label {
                    display: block;
                    font-weight: 600;
                    color: #555;
                    margin-bottom: 8px;
                }
                input[type="file"] {
                    width: 100%;
                    padding: 12px;
                    border: 2px dashed #ddd;
                    border-radius: 8px;
                    background: #f8f9fa;
                    cursor: pointer;
                    transition: all 0.3s;
                }
                input[type="file"]:hover {
                    border-color: #667eea;
                    background: #f0f0ff;
                }
                button {
                    width: 100%;
                    padding: 15px;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    border: none;
                    border-radius: 8px;
                    font-size: 16px;
                    font-weight: 600;
                    cursor: pointer;
                    transition: transform 0.2s;
                }
                button:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
                }
                button:active {
                    transform: translateY(0);
                }
                .footer {
                    margin-top: 30px;
                    text-align: center;
                    color: #999;
                    font-size: 13px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🔄 Restaurar Backup</h1>
                <p style="color: #666; margin-top: 5px;">Sube un archivo ZIP con productos.db e imágenes</p>
                
                <div class="warning">
                    <strong>⚠️ ADVERTENCIA:</strong> Esta operación reemplazará completamente la base de datos y las imágenes actuales. No se puede deshacer.
                </div>
                
                <div class="info">
                    <strong>📦 El archivo ZIP debe contener:</strong>
                    <ul>
                        <li><code>productos.db</code> (base de datos)</li>
                        <li><code>img/</code> (carpeta con imágenes)</li>
                    </ul>
                    <p style="margin-top: 10px;"><strong>Se extraerá a:</strong></p>
                    <ul>
                        <li>→ <code>/data/productos.db</code></li>
                        <li>→ <code>/data/img/</code></li>
                    </ul>
                </div>
                
                <form method="POST" enctype="multipart/form-data">
                    <div class="form-group">
                        <label for="backup_zip">Seleccionar archivo ZIP:</label>
                        <input type="file" id="backup_zip" name="backup_zip" accept=".zip" required>
                    </div>
                    
                    <button type="submit">🚀 Restaurar Backup</button>
                </form>
                
                <div class="footer">
                    RM KITS - Sistema de Restauración Temporal
                </div>
            </div>
        </body>
        </html>
        """
        return html
    
    # POST: Procesar el archivo ZIP
    try:
        # Validar que se haya subido un archivo
        if 'backup_zip' not in request.files:
            return """
            <html>
            <body style="font-family: Arial; padding: 40px; text-align: center;">
                <h2 style="color: #dc3545;">❌ Error</h2>
                <p>No se seleccionó ningún archivo</p>
                <a href="/restaurar" style="display: inline-block; margin-top: 20px; padding: 10px 20px; background: #667eea; color: white; text-decoration: none; border-radius: 5px;">← Volver</a>
            </body>
            </html>
            """, 400
        
        file = request.files['backup_zip']
        
        # Validar que el archivo tenga nombre
        if file.filename == '':
            return """
            <html>
            <body style="font-family: Arial; padding: 40px; text-align: center;">
                <h2 style="color: #dc3545;">❌ Error</h2>
                <p>No se seleccionó ningún archivo</p>
                <a href="/restaurar" style="display: inline-block; margin-top: 20px; padding: 10px 20px; background: #667eea; color: white; text-decoration: none; border-radius: 5px;">← Volver</a>
            </body>
            </html>
            """, 400
        
        # Validar extensión .zip
        if not file.filename.lower().endswith('.zip'):
            return """
            <html>
            <body style="font-family: Arial; padding: 40px; text-align: center;">
                <h2 style="color: #dc3545;">❌ Error</h2>
                <p>El archivo debe ser un ZIP (.zip)</p>
                <a href="/restaurar" style="display: inline-block; margin-top: 20px; padding: 10px 20px; background: #667eea; color: white; text-decoration: none; border-radius: 5px;">← Volver</a>
            </body>
            </html>
            """, 400
        
        # Leer el ZIP en memoria
        logger.info(f"📦 Procesando archivo: {file.filename}")
        zip_data = BytesIO(file.read())
        
        # Contadores
        db_restaurada = False
        imagenes_restauradas = 0
        archivos_ignorados = 0
        
        # Crear carpeta de imágenes si no existe
        os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
        
        # Extraer archivos del ZIP
        with zipfile.ZipFile(zip_data, 'r') as zip_ref:
            archivos_zip = zip_ref.namelist()
            logger.info(f"Archivos en el ZIP: {len(archivos_zip)}")
            
            for archivo in archivos_zip:
                # Ignorar carpetas vacías
                if archivo.endswith('/'):
                    continue
                
                # Obtener el nombre base del archivo (sin rutas)
                nombre_archivo = os.path.basename(archivo)
                
                # Restaurar productos.db
                if nombre_archivo == 'productos.db' or archivo.endswith('/productos.db'):
                    logger.info(f"✓ Extrayendo base de datos: {archivo}")
                    with zip_ref.open(archivo) as source:
                        with open(Config.DATABASE_PATH, 'wb') as target:
                            shutil.copyfileobj(source, target)
                    db_restaurada = True
                    logger.info(f"✅ Base de datos restaurada en: {Config.DATABASE_PATH}")
                
                # Restaurar imágenes de la carpeta img/
                elif '/img/' in archivo or archivo.startswith('img/'):
                    if nombre_archivo:  # Asegurar que no sea una carpeta vacía
                        logger.info(f"✓ Extrayendo imagen: {nombre_archivo}")
                        destino = os.path.join(Config.UPLOAD_FOLDER, nombre_archivo)
                        with zip_ref.open(archivo) as source:
                            with open(destino, 'wb') as target:
                                shutil.copyfileobj(source, target)
                        imagenes_restauradas += 1
                
                else:
                    archivos_ignorados += 1
                    logger.info(f"ℹ️ Ignorado: {archivo}")
        
        # Generar resultado
        resultado_html = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Restauración Completada</title>
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    padding: 20px;
                }}
                .container {{
                    background: white;
                    border-radius: 16px;
                    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                    padding: 40px;
                    max-width: 600px;
                    width: 100%;
                }}
                h1 {{
                    color: #28a745;
                    margin-bottom: 20px;
                    font-size: 32px;
                }}
                .stats {{
                    background: #f8f9fa;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 20px 0;
                }}
                .stat {{
                    display: flex;
                    justify-content: space-between;
                    padding: 10px 0;
                    border-bottom: 1px solid #e0e0e0;
                }}
                .stat:last-child {{
                    border-bottom: none;
                }}
                .label {{
                    font-weight: 600;
                    color: #555;
                }}
                .value {{
                    color: #667eea;
                    font-weight: 700;
                }}
                .success {{
                    background: #d4edda;
                    color: #155724;
                    padding: 15px;
                    border-radius: 8px;
                    border-left: 4px solid #28a745;
                    margin: 20px 0;
                }}
                .warning {{
                    background: #fff3cd;
                    color: #856404;
                    padding: 15px;
                    border-radius: 8px;
                    border-left: 4px solid #ffc107;
                    margin: 20px 0;
                }}
                .button {{
                    display: inline-block;
                    margin: 10px 5px;
                    padding: 12px 24px;
                    background: #667eea;
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    font-weight: 600;
                    transition: transform 0.2s;
                }}
                .button:hover {{
                    transform: translateY(-2px);
                }}
                .button.secondary {{
                    background: #6c757d;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>✅ Restauración Completada</h1>
                
                <div class="stats">
                    <div class="stat">
                        <span class="label">📄 Base de datos:</span>
                        <span class="value">{'✓ Restaurada' if db_restaurada else '✗ No encontrada'}</span>
                    </div>
                    <div class="stat">
                        <span class="label">🖼️ Imágenes restauradas:</span>
                        <span class="value">{imagenes_restauradas}</span>
                    </div>
                    <div class="stat">
                        <span class="label">⏭️ Archivos ignorados:</span>
                        <span class="value">{archivos_ignorados}</span>
                    </div>
                </div>
                
                {'<div class="success"><strong>✓ Base de datos restaurada correctamente</strong><br>Ubicación: ' + Config.DATABASE_PATH + '</div>' if db_restaurada else '<div class="warning"><strong>⚠️ No se encontró productos.db en el ZIP</strong></div>'}
                
                {'<div class="success"><strong>✓ Imágenes restauradas</strong><br>Ubicación: ' + Config.UPLOAD_FOLDER + '</div>' if imagenes_restauradas > 0 else ''}
                
                <div style="text-align: center; margin-top: 30px;">
                    <a href="/" class="button">🏠 Ir al inicio</a>
                    <a href="/admin/login" class="button secondary">🔐 Panel Admin</a>
                </div>
                
                <div style="margin-top: 30px; padding: 15px; background: #fff3cd; border-radius: 8px; font-size: 14px; color: #856404;">
                    <strong>⚠️ Recordatorio:</strong> Elimina la ruta /restaurar del código después de usar.
                </div>
            </div>
        </body>
        </html>
        """
        
        return resultado_html
        
    except zipfile.BadZipFile:
        logger.error("❌ Archivo ZIP corrupto o inválido")
        return """
        <html>
        <body style="font-family: Arial; padding: 40px; text-align: center;">
            <h2 style="color: #dc3545;">❌ Error</h2>
            <p>El archivo no es un ZIP válido o está corrupto</p>
            <a href="/restaurar" style="display: inline-block; margin-top: 20px; padding: 10px 20px; background: #667eea; color: white; text-decoration: none; border-radius: 5px;">← Volver</a>
        </body>
        </html>
        """, 400
    except Exception as e:
        logger.error(f"❌ Error al restaurar backup: {e}")
        return f"""
        <html>
        <body style="font-family: Arial; padding: 40px; text-align: center;">
            <h2 style="color: #dc3545;">❌ Error</h2>
            <p>Error al procesar el backup: {str(e)}</p>
            <a href="/restaurar" style="display: inline-block; margin-top: 20px; padding: 10px 20px; background: #667eea; color: white; text-decoration: none; border-radius: 5px;">← Volver</a>
        </body>
        </html>
        """, 500


# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================

def buscar_imagen_para_codigo(codigo):
    """
    Busca si existe una imagen para el código dado en el almacenamiento persistente.
    Retorna el nombre del archivo si existe, o string vacío si no.
    """
    extensiones = ['.jpg', '.jpeg', '.png', '.webp', '.gif']
    img_folder = Config.UPLOAD_FOLDER
    
    for ext in extensiones:
        img_filename = f"{codigo}{ext}"
        img_path = os.path.join(img_folder, img_filename)
        if os.path.exists(img_path):
            return img_filename
    
    return ""


def generar_codigo_producto():
    """Genera el siguiente código de producto automáticamente (formato A0XXX)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            # Obtener el último código que empiece con 'A0'
            cursor.execute("""
                SELECT codigo FROM producto 
                WHERE codigo LIKE 'A0%' 
                ORDER BY CAST(SUBSTR(codigo, 2) AS INTEGER) DESC 
                LIMIT 1
            """)
            result = cursor.fetchone()
            
            if result:
                ultimo_codigo = result['codigo']
                # Extraer el número del código (A0222 -> 222)
                numero = int(ultimo_codigo[1:])
                nuevo_numero = numero + 1
            else:
                # Si no hay códigos, empezar desde 1
                nuevo_numero = 1
            
            # Formatear el nuevo código (A0001, A0222, etc)
            nuevo_codigo = f"A{nuevo_numero:04d}"
            return nuevo_codigo
    except Exception as e:
        logger.error(f"Error al generar código: {e}")
        return "A0001"


def init_database():
    """Inicializa las tablas necesarias en la base de datos"""
    try:
        # Verificar que la DB ya existe antes de conectar
        # (sqlite3.connect crea una DB nueva si no existe, lo cual queremos evitar)
        if not os.path.exists(Config.DATABASE_PATH):
            logger.error(f"❌ No se puede inicializar: la base de datos no existe en {Config.DATABASE_PATH}")
            return
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Verificar que la tabla producto existe (asegurar que no es una DB vacía nueva)
            cursor.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='producto'
            """)
            if not cursor.fetchone():
                logger.error("❌ ERROR: La base de datos no contiene la tabla 'producto'")
                logger.error("❌ Parece que la base de datos está vacía o corrupta")
                raise Exception("Base de datos inválida: falta la tabla 'producto'")
            
            logger.info("✓ Tabla 'producto' encontrada en la base de datos")
            
            # Crear tabla categorías si no existe
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS categoria (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL UNIQUE,
                    fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Crear tabla productos_nuevos si no existe
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS producto_nuevo (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    producto_id INTEGER NOT NULL,
                    fecha_agregado TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (producto_id) REFERENCES producto(id)
                )
            """)
            
            # Crear tabla pedidos si no existe
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS pedido (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    cliente_nombre TEXT NOT NULL,
                    cliente_cuit TEXT,
                    cliente_telefono TEXT,
                    cliente_email TEXT,
                    cliente_direccion TEXT,
                    metodo_entrega TEXT,
                    envio_direccion TEXT,
                    envio_localidad TEXT,
                    envio_provincia TEXT,
                    envio_cp TEXT,
                    envio_nombre_destinatario TEXT,
                    envio_dni_destinatario TEXT,
                    envio_referencias TEXT,
                    productos TEXT NOT NULL,
                    total REAL NOT NULL,
                    estado TEXT DEFAULT 'pendiente'
                )
            """)

            try:
                cursor.execute("SELECT envio_dni_destinatario FROM pedido LIMIT 1")
            except sqlite3.OperationalError:
                cursor.execute("ALTER TABLE pedido ADD COLUMN envio_dni_destinatario TEXT")
            
            # Verificar si hay pedidos existentes
            cursor.execute("SELECT COUNT(*) as count FROM pedido")
            count = cursor.fetchone()[0]
            
            # Si no hay pedidos, configurar el AUTOINCREMENT para empezar en 1200
            if count == 0:
                cursor.execute("INSERT OR REPLACE INTO sqlite_sequence (name, seq) VALUES ('pedido', 1199)")
            else:
                # Si ya hay pedidos, verificar que el próximo ID sea al menos 1200
                cursor.execute("SELECT MAX(id) as max_id FROM pedido")
                max_id = cursor.fetchone()[0]
                if max_id is None or max_id < 1199:
                    cursor.execute("INSERT OR REPLACE INTO sqlite_sequence (name, seq) VALUES ('pedido', 1199)")
            
            conn.commit()
            logger.info("✓ Tablas de base de datos inicializadas correctamente")
    except FileNotFoundError:
        raise
    except Exception as e:
        logger.error(f"❌ Error al inicializar base de datos: {e}")
        raise


def migrar_categorias_existentes():
    """Migra las categorías existentes de los productos a la tabla de categorías"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Obtener categorías únicas de los productos
            cursor.execute("SELECT DISTINCT categoria FROM producto WHERE categoria IS NOT NULL AND categoria != '' ORDER BY categoria")
            categorias_existentes = [row['categoria'] for row in cursor.fetchall()]
            
            # Insertar categorías en la tabla de categorías si no existen
            for categoria in categorias_existentes:
                cursor.execute("INSERT OR IGNORE INTO categoria (nombre) VALUES (?)", (categoria,))
            
            conn.commit()
            logger.info(f"Migradas {len(categorias_existentes)} categorías existentes")
    except Exception as e:
        logger.error(f"Error al migrar categorías existentes: {e}")


# Inicializar base de datos al arrancar la app
init_database()
migrar_categorias_existentes()

# =============================================================================
# PANEL DE ADMINISTRACIÓN
# =============================================================================

def login_required(f):
    """Decorador para requerir login en rutas de admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Debes iniciar sesión para acceder al panel de administración', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def allowed_file(filename):
    """Verifica si el archivo tiene una extensión permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    """Login del administrador"""
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        
        # Verificar si el usuario existe y la contraseña es correcta
        if username in Config.ADMIN_USERS and Config.ADMIN_USERS[username] == password:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            flash('Inicio de sesión exitoso', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template("admin/login.html")


@app.route("/admin/logout")
def admin_logout():
    """Cerrar sesión del administrador"""
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    flash('Sesión cerrada exitosamente', 'success')
    return redirect(url_for('admin_login'))


@app.route("/admin")
@login_required
def admin_dashboard():
    """Dashboard principal del admin"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Estadísticas de productos
            cursor.execute("SELECT COUNT(*) FROM producto")
            total_productos = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM producto WHERE stock > 0")
            productos_con_stock = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM producto WHERE stock = 0")
            productos_sin_stock = cursor.fetchone()[0]
            
            cursor.execute("SELECT SUM(stock) FROM producto")
            stock_total = cursor.fetchone()[0] or 0
            
            # Estadísticas de pedidos
            cursor.execute("SELECT COUNT(*) FROM pedido")
            total_pedidos = cursor.fetchone()[0]
            
            cursor.execute("SELECT SUM(total) FROM pedido")
            total_ventas = cursor.fetchone()[0] or 0
            
            # Ventas facturadas vs por facturar
            cursor.execute("""
                SELECT SUM(total) FROM pedido 
                WHERE estado IN ('pagado', 'completado', 'impreso', 'preparado')
            """)
            ventas_facturadas = cursor.fetchone()[0] or 0
            
            cursor.execute("""
                SELECT SUM(total) FROM pedido 
                WHERE estado NOT IN ('pagado', 'completado', 'impreso', 'preparado', 'cancelado')
            """)
            ventas_por_facturar = cursor.fetchone()[0] or 0
            
            # Ventas por día (últimos 30 días)
            cursor.execute("""
                SELECT DATE(fecha) as dia, COUNT(*) as cantidad, SUM(total) as monto
                FROM pedido
                WHERE fecha >= date('now', '-30 days')
                GROUP BY DATE(fecha)
                ORDER BY dia DESC
            """)
            ventas_por_dia_raw = cursor.fetchall()
            ventas_por_dia = [list(row) for row in ventas_por_dia_raw] if ventas_por_dia_raw else []
            
            # Productos más vendidos
            cursor.execute("""
                SELECT p.productos, COUNT(*) as veces_pedido
                FROM pedido p
                WHERE p.productos IS NOT NULL AND p.productos != ''
                GROUP BY p.productos
                ORDER BY veces_pedido DESC
                LIMIT 10
            """)
            productos_vendidos_raw = cursor.fetchall()
            
            # Procesar productos más vendidos
            from collections import Counter
            productos_counter = Counter()
            
            for row in productos_vendidos_raw:
                try:
                    productos_json = json.loads(row[0])
                    for prod in productos_json:
                        titulo = prod.get('titulo', 'Sin nombre')
                        cantidad = prod.get('cantidad', 0)
                        productos_counter[titulo] += cantidad
                except Exception as e:
                    logger.error(f"Error procesando producto: {e}")
                    pass
            
            productos_mas_vendidos = list(productos_counter.most_common(10)) if productos_counter else []
            
            return render_template("admin/dashboard.html",
                                 total_productos=total_productos,
                                 productos_con_stock=productos_con_stock,
                                 productos_sin_stock=productos_sin_stock,
                                 stock_total=stock_total,
                                 total_pedidos=total_pedidos)
    except Exception as e:
        logger.error(f"Error en dashboard: {e}")
        flash('Error al cargar el dashboard', 'error')
        return redirect(url_for('admin_login'))


@app.route("/admin/api/ventas-por-dia")
@login_required
def admin_api_ventas_por_dia():
    """API para obtener ventas por día según período"""
    try:
        periodo = request.args.get('periodo', '30')
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Construir query según período
            if periodo == 'all':
                query = """
                    SELECT DATE(fecha) as dia, COUNT(*) as cantidad, SUM(total) as monto
                    FROM pedido
                    GROUP BY DATE(fecha)
                    ORDER BY dia ASC
                """
            else:
                dias = int(periodo)
                query = f"""
                    SELECT DATE(fecha) as dia, COUNT(*) as cantidad, SUM(total) as monto
                    FROM pedido
                    WHERE fecha >= date('now', '-{dias} days')
                    GROUP BY DATE(fecha)
                    ORDER BY dia ASC
                """
            
            cursor.execute(query)
            ventas_raw = cursor.fetchall()
            ventas = [list(row) for row in ventas_raw] if ventas_raw else []
            
            return jsonify(ventas)
    except Exception as e:
        logger.error(f"Error en API ventas por día: {e}")
        return jsonify([]), 500


@app.route("/admin/api/productos-mas-vendidos")
@login_required
def admin_api_productos_mas_vendidos():
    """API para obtener productos más vendidos con filtro opcional"""
    try:
        filtro = request.args.get('filtro', '').lower()
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Obtener todos los pedidos
            cursor.execute("""
                SELECT p.productos
                FROM pedido p
                WHERE p.productos IS NOT NULL AND p.productos != ''
            """)
            productos_vendidos_raw = cursor.fetchall()
            
            # Procesar productos más vendidos con filtro
            from collections import Counter
            productos_counter = Counter()
            
            for row in productos_vendidos_raw:
                try:
                    productos_json = json.loads(row[0])
                    for prod in productos_json:
                        titulo = prod.get('titulo', 'Sin nombre')
                        # Aplicar filtro si existe
                        if not filtro or filtro in titulo.lower():
                            cantidad = prod.get('cantidad', 0)
                            productos_counter[titulo] += cantidad
                except Exception as e:
                    logger.error(f"Error procesando producto: {e}")
                    pass
            
            productos_mas_vendidos = list(productos_counter.most_common(10)) if productos_counter else []
            
            return jsonify(productos_mas_vendidos)
    except Exception as e:
        logger.error(f"Error en API productos más vendidos: {e}")
        return jsonify([]), 500


def _normalizar_ruta_backup(path):
    """Normaliza rutas de archivos de backup para lectura robusta"""
    return (path or '').replace('\\', '/').lstrip('./')


def _buscar_archivo_backup(archivos_map, filename):
    """Busca un archivo por nombre dentro de un mapa ruta->bytes"""
    objetivo = filename.lower()
    for ruta, contenido in archivos_map.items():
        ruta_norm = _normalizar_ruta_backup(ruta).lower()
        if ruta_norm == objetivo or ruta_norm.endswith('/' + objetivo):
            return contenido
    return None


def _cargar_archivos_backup_desde_request():
    """Carga archivos de backup desde ZIP o carpeta subida por el navegador"""
    archivos_map = {}

    archivo_zip = request.files.get('archivo')
    if archivo_zip and archivo_zip.filename:
        try:
            zip_bytes = archivo_zip.read()
            with zipfile.ZipFile(BytesIO(zip_bytes), 'r') as zip_file:
                for info in zip_file.infolist():
                    if info.is_dir():
                        continue
                    archivos_map[_normalizar_ruta_backup(info.filename)] = zip_file.read(info.filename)
            return archivos_map
        except Exception as e:
            raise ValueError(f"El ZIP no es válido: {e}")

    archivos = request.files.getlist('archivos')
    for file in archivos:
        if not file or not file.filename:
            continue
        ruta = _normalizar_ruta_backup(file.filename)
        if not ruta:
            continue
        archivos_map[ruta] = file.read()

    if archivos_map:
        return archivos_map

    raise ValueError("No se recibió ningún archivo para importar")


@app.route("/admin/exportar-todo")
@login_required
def admin_exportar_todo():
    """Exporta productos, pedidos, productos_nuevos e imágenes de productos nuevos"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM producto ORDER BY codigo DESC")
            productos = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT * FROM pedido ORDER BY id")
            pedidos = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT * FROM producto_nuevo ORDER BY id")
            productos_nuevos = [dict(row) for row in cursor.fetchall()]

            cursor.execute("""
                SELECT DISTINCT p.imagen
                FROM producto p
                INNER JOIN producto_nuevo pn ON p.id = pn.producto_id
                WHERE p.imagen IS NOT NULL AND p.imagen != ''
            """)
            imagenes_nuevos = [row['imagen'] for row in cursor.fetchall()]

        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        carpeta_base = f"backup_rmkits_{fecha_actual}"

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            metadata = {
                'app': 'RM KITS',
                'version': 1,
                'exportado_en': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'productos': len(productos),
                'pedidos': len(pedidos),
                'productos_nuevos': len(productos_nuevos),
                'imagenes_nuevos': len(imagenes_nuevos)
            }

            zip_file.writestr(
                f"{carpeta_base}/metadata.json",
                json.dumps(metadata, ensure_ascii=False, indent=2)
            )
            zip_file.writestr(
                f"{carpeta_base}/productos.json",
                json.dumps(productos, ensure_ascii=False, indent=2)
            )
            zip_file.writestr(
                f"{carpeta_base}/pedidos.json",
                json.dumps(pedidos, ensure_ascii=False, indent=2)
            )
            zip_file.writestr(
                f"{carpeta_base}/productos_nuevos.json",
                json.dumps(productos_nuevos, ensure_ascii=False, indent=2)
            )

            for imagen in imagenes_nuevos:
                imagen_path = os.path.join(app.config['UPLOAD_FOLDER'], imagen)
                if not os.path.exists(imagen_path):
                    continue

                nombre_archivo = secure_filename(os.path.basename(imagen))
                if not nombre_archivo:
                    continue

                zip_file.write(
                    imagen_path,
                    arcname=f"{carpeta_base}/imagenes_nuevos/{nombre_archivo}"
                )

        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"backup_rmkits_{fecha_actual}.zip"
        )

    except Exception as e:
        logger.error(f"Error al exportar todo: {e}")
        flash(f'Error al exportar todo: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route("/admin/importar-todo", methods=["POST"])
@login_required
def admin_importar_todo():
    """Importa backup completo de productos, pedidos, productos_nuevos e imágenes"""
    try:
        archivos_map = _cargar_archivos_backup_desde_request()

        productos_bytes = _buscar_archivo_backup(archivos_map, 'productos.json')
        pedidos_bytes = _buscar_archivo_backup(archivos_map, 'pedidos.json')
        productos_nuevos_bytes = _buscar_archivo_backup(archivos_map, 'productos_nuevos.json')

        if not productos_bytes or not pedidos_bytes or not productos_nuevos_bytes:
            return jsonify({
                'success': False,
                'error': 'Faltan archivos requeridos: productos.json, pedidos.json y productos_nuevos.json'
            }), 400

        try:
            productos = json.loads(productos_bytes.decode('utf-8-sig'))
            pedidos = json.loads(pedidos_bytes.decode('utf-8-sig'))
            productos_nuevos = json.loads(productos_nuevos_bytes.decode('utf-8-sig'))
        except Exception as e:
            return jsonify({'success': False, 'error': f'JSON inválido en backup: {str(e)}'}), 400

        if not isinstance(productos, list) or not isinstance(pedidos, list) or not isinstance(productos_nuevos, list):
            return jsonify({'success': False, 'error': 'El formato del backup es inválido'}), 400

        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Asegurar columnas actuales (migración defensiva)
            try:
                cursor.execute("SELECT activo FROM producto LIMIT 1")
            except sqlite3.OperationalError:
                cursor.execute("ALTER TABLE producto ADD COLUMN activo INTEGER NOT NULL DEFAULT 1")

            columnas_pedido = [
                ('envio_direccion', 'TEXT'),
                ('envio_localidad', 'TEXT'),
                ('envio_provincia', 'TEXT'),
                ('envio_cp', 'TEXT'),
                ('envio_nombre_destinatario', 'TEXT'),
                ('envio_dni_destinatario', 'TEXT'),
                ('envio_referencias', 'TEXT')
            ]
            for columna, tipo in columnas_pedido:
                try:
                    cursor.execute(f"SELECT {columna} FROM pedido LIMIT 1")
                except sqlite3.OperationalError:
                    cursor.execute(f"ALTER TABLE pedido ADD COLUMN {columna} {tipo}")

            # Reemplazo total de datos
            cursor.execute("DELETE FROM producto_nuevo")
            cursor.execute("DELETE FROM pedido")
            cursor.execute("DELETE FROM producto")

            for p in productos:
                cursor.execute("""
                    INSERT INTO producto (
                        id, codigo, titulo, descripcion, precio, minimo,
                        multiplo, stock, imagen, categoria, activo
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    p.get('id'),
                    p.get('codigo', ''),
                    p.get('titulo', ''),
                    p.get('descripcion', ''),
                    p.get('precio', 0),
                    p.get('minimo', 1),
                    p.get('multiplo', 1),
                    p.get('stock', 0),
                    p.get('imagen', ''),
                    p.get('categoria', ''),
                    p.get('activo', 1)
                ))

            for ped in pedidos:
                cursor.execute("""
                    INSERT INTO pedido (
                        id, fecha, cliente_nombre, cliente_cuit, cliente_telefono,
                        cliente_email, cliente_direccion, metodo_entrega,
                        envio_direccion, envio_localidad, envio_provincia, envio_cp,
                        envio_nombre_destinatario, envio_dni_destinatario, envio_referencias,
                        productos, total, estado
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    ped.get('id'),
                    ped.get('fecha'),
                    ped.get('cliente_nombre', ''),
                    ped.get('cliente_cuit', ''),
                    ped.get('cliente_telefono', ''),
                    ped.get('cliente_email', ''),
                    ped.get('cliente_direccion', ''),
                    ped.get('metodo_entrega', 'retiro'),
                    ped.get('envio_direccion', ''),
                    ped.get('envio_localidad', ''),
                    ped.get('envio_provincia', ''),
                    ped.get('envio_cp', ''),
                    ped.get('envio_nombre_destinatario', ''),
                    ped.get('envio_dni_destinatario', ped.get('envio_cuit_destinatario', '')),
                    ped.get('envio_referencias', ''),
                    ped.get('productos', '[]'),
                    ped.get('total', 0),
                    ped.get('estado', 'pendiente')
                ))

            for pn in productos_nuevos:
                cursor.execute("""
                    INSERT INTO producto_nuevo (id, producto_id, fecha_agregado)
                    VALUES (?, ?, ?)
                """, (
                    pn.get('id'),
                    pn.get('producto_id'),
                    pn.get('fecha_agregado')
                ))

            # Ajustar autoincrementos
            for tabla in ['producto', 'pedido', 'producto_nuevo']:
                cursor.execute(f"SELECT MAX(id) FROM {tabla}")
                max_id = cursor.fetchone()[0]
                cursor.execute("DELETE FROM sqlite_sequence WHERE name = ?", (tabla,))
                if max_id is not None:
                    cursor.execute("INSERT INTO sqlite_sequence (name, seq) VALUES (?, ?)", (tabla, max_id))

            conn.commit()

        # Importar imágenes de productos nuevos (sin borrar imágenes existentes)
        imagenes_importadas = 0
        for ruta, contenido in archivos_map.items():
            ruta_norm = _normalizar_ruta_backup(ruta).lower()
            if '/imagenes_nuevos/' not in ruta_norm and not ruta_norm.startswith('imagenes_nuevos/'):
                continue

            nombre_archivo = secure_filename(os.path.basename(ruta))
            if not nombre_archivo:
                continue

            destino = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
            with open(destino, 'wb') as f:
                f.write(contenido)
            imagenes_importadas += 1

        return jsonify({
            'success': True,
            'productos': len(productos),
            'pedidos': len(pedidos),
            'productos_nuevos': len(productos_nuevos),
            'imagenes': imagenes_importadas
        })

    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Error al importar todo: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/admin/productos")
@login_required
def admin_productos():
    """Lista de productos en el admin"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM producto ORDER BY codigo DESC")
            productos = [dict(row) for row in cursor.fetchall()]
            
            # Obtener categorías de la tabla de categorías
            cursor.execute("SELECT nombre FROM categoria ORDER BY nombre")
            categorias = [row['nombre'] for row in cursor.fetchall()]
            
            return render_template("admin/productos.html", productos=productos, categorias=categorias)
    except Exception as e:
        logger.error(f"Error al obtener productos: {e}")
        flash('Error al cargar productos', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route("/admin/descargar-excel")
@login_required
def admin_descargar_excel():
    """Descargar lista de productos como Excel"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM producto ORDER BY codigo DESC")
            productos = [dict(row) for row in cursor.fetchall()]
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Estilos
        header_fill = PatternFill(start_color="6a1b9a", end_color="6a1b9a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ["ID", "Código", "Título", "Descripción", "Precio", "Mínimo", "Múltiplo", "Stock", "Categoría", "Activo"]
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Agregar datos
        for producto in productos:
            ws.append([
                producto.get('id', ''),
                producto.get('codigo', ''),
                producto.get('titulo', ''),
                producto.get('descripcion', ''),
                producto.get('precio', 0),
                producto.get('minimo', 1),
                producto.get('multiplo', 1),
                producto.get('stock', 0),
                producto.get('categoria', ''),
                'Sí' if producto.get('activo', 1) == 1 else 'No'
            ])
        
        # Ajustar ancho de columnas
        column_widths = [8, 15, 40, 50, 12, 10, 10, 10, 20, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        filename = f"productos_rmkits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Error al generar Excel: {e}")
        flash('Error al generar archivo Excel', 'error')
        return redirect(url_for('admin_productos'))


@app.route("/admin/subir-excel", methods=["POST"])
@login_required
def admin_subir_excel():
    """Subir archivo Excel para actualizar productos"""
    try:
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('admin_productos'))
        
        file = request.files['archivo']
        
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('admin_productos'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('El archivo debe ser un Excel (.xlsx o .xls)', 'error')
            return redirect(url_for('admin_productos'))
        
        # Leer el archivo Excel
        wb = load_workbook(file)
        ws = wb.active
        
        # Verificar encabezados
        headers = [cell.value for cell in ws[1]]
        expected_headers = ["ID", "Código", "Título", "Descripción", "Precio", "Mínimo", "Múltiplo", "Stock", "Categoría", "Activo"]
        
        if headers != expected_headers:
            flash(f'El formato del archivo no es correcto. Se esperan las columnas: {", ".join(expected_headers)}', 'error')
            return redirect(url_for('admin_productos'))
        
        # Procesar filas
        productos_actualizados = 0
        productos_creados = 0
        errores = []
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    id_producto = row[0]
                    codigo = row[1]
                    titulo = row[2]
                    descripcion = row[3] or ''
                    precio = float(row[4]) if row[4] else 0
                    minimo = int(row[5]) if row[5] else 1
                    multiplo = int(row[6]) if row[6] else 1
                    stock = int(row[7]) if row[7] else 0
                    categoria = row[8] or ''
                    activo = 1 if row[9] in ['Sí', 'Si', 'SI', 'SÍ', 1, '1', True] else 0
                    
                    # Verificar si el producto existe por ID o código
                    if id_producto:
                        cursor.execute("SELECT id FROM producto WHERE id = ?", (id_producto,))
                        existe = cursor.fetchone()
                        
                        if existe:
                            # Actualizar producto existente
                            cursor.execute("""
                                UPDATE producto 
                                SET codigo=?, titulo=?, descripcion=?, precio=?, minimo=?, multiplo=?, stock=?, categoria=?, activo=?
                                WHERE id=?
                            """, (codigo, titulo, descripcion, precio, minimo, multiplo, stock, categoria, activo, id_producto))
                            productos_actualizados += 1
                        else:
                            # Crear nuevo producto con ID específico
                            # Buscar imagen automáticamente si no está en el Excel
                            imagen_auto = buscar_imagen_para_codigo(codigo)
                            cursor.execute("""
                                INSERT INTO producto (id, codigo, titulo, descripcion, precio, minimo, multiplo, stock, imagen, categoria, activo)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (id_producto, codigo, titulo, descripcion, precio, minimo, multiplo, stock, imagen_auto, categoria, activo))
                            productos_creados += 1
                    else:
                        # Buscar por código si no hay ID
                        cursor.execute("SELECT id FROM producto WHERE codigo = ?", (codigo,))
                        existe = cursor.fetchone()
                        
                        if existe:
                            # Actualizar producto existente
                            cursor.execute("""
                                UPDATE producto 
                                SET titulo=?, descripcion=?, precio=?, minimo=?, multiplo=?, stock=?, categoria=?, activo=?
                                WHERE codigo=?
                            """, (titulo, descripcion, precio, minimo, multiplo, stock, categoria, activo, codigo))
                            productos_actualizados += 1
                        else:
                            # Crear nuevo producto
                            # Buscar imagen automáticamente si no está en el Excel
                            imagen_auto = buscar_imagen_para_codigo(codigo)
                            cursor.execute("""
                                INSERT INTO producto (codigo, titulo, descripcion, precio, minimo, multiplo, stock, imagen, categoria, activo)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (codigo, titulo, descripcion, precio, minimo, multiplo, stock, imagen_auto, categoria, activo))
                            productos_creados += 1
                
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
            
            conn.commit()
        
        # Mensaje de resultado
        mensaje = f'✅ Procesamiento completado: {productos_actualizados} actualizados, {productos_creados} creados'
        if errores:
            mensaje += f'. ⚠️ {len(errores)} errores encontrados'
            logger.warning(f"Errores al importar Excel: {errores}")
        
        flash(mensaje, 'success' if not errores else 'warning')
        return redirect(url_for('admin_productos'))
    
    except Exception as e:
        logger.error(f"Error al procesar Excel: {e}")
        flash(f'Error al procesar archivo: {str(e)}', 'error')
        return redirect(url_for('admin_productos'))


@app.route("/admin/pedidos")
@login_required
def admin_pedidos():
    """Ver lista de pedidos"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM pedido ORDER BY fecha DESC")
            pedidos = [dict(row) for row in cursor.fetchall()]
            
            return render_template("admin/pedidos.html", pedidos=pedidos)
    except Exception as e:
        logger.error(f"Error al obtener pedidos: {e}")
        flash('Error al cargar pedidos', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route("/admin/cargar-pedido-manual", methods=['POST'])
@login_required
def admin_cargar_pedido_manual():
    """Cargar pedido manualmente desde datos estructurados"""
    try:
        data = request.json
        
        # Preparar productos para JSON
        productos_json = json.dumps(data['productos'])
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO pedido (cliente_nombre, cliente_telefono, cliente_email,
                                   metodo_entrega, envio_direccion, envio_localidad,
                                   envio_provincia, envio_cp, envio_nombre_destinatario,
                                   productos, total, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data['cliente_nombre'],
                data['cliente_telefono'],
                data.get('cliente_email', ''),
                data['metodo_entrega'],
                data.get('envio_direccion', ''),
                data.get('envio_localidad', ''),
                data.get('envio_provincia', ''),
                data.get('envio_cp', ''),
                data.get('envio_nombre_destinatario', ''),
                productos_json,
                data['total'],
                data.get('estado', 'pendiente')
            ))
            
            pedido_id = cursor.lastrowid
            conn.commit()
        
        return jsonify({'success': True, 'pedido_id': pedido_id}), 200
    
    except Exception as e:
        logger.error(f"Error al cargar pedido manual: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
        logger.error(f"Error al obtener pedidos: {e}")
        flash('Error al cargar pedidos', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route("/admin/pedidos/exportar")
@login_required
def admin_exportar_pedidos():
    """Exportar pedidos a Excel - Genera 2 archivos en un ZIP"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM pedido ORDER BY fecha DESC")
            pedidos = [dict(row) for row in cursor.fetchall()]
        
        # Estilos comunes
        header_fill = PatternFill(start_color="6a1b9a", end_color="6a1b9a", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # ===== ARCHIVO 1: DATOS DEL PEDIDO =====
        wb_datos = Workbook()
        ws_datos = wb_datos.active
        ws_datos.title = "Datos Pedidos"
        
        # Encabezados para datos
        headers = [
            "ID", "Fecha", "Cliente", "CUIT", "Teléfono", "Email", 
            "Método Entrega", "Dirección Envío", "Localidad", "Provincia", 
            "CP", "Destinatario", "Referencias", "Total", "Estado"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_datos.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos de pedidos
        for row_num, pedido in enumerate(pedidos, 2):
            ws_datos.cell(row=row_num, column=1, value=pedido.get('id', ''))
            
            # Formatear fecha
            fecha = pedido.get('fecha', '')
            if fecha:
                try:
                    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d %H:%M:%S')
                    fecha = fecha_obj.strftime('%d/%m/%Y %H:%M')
                except:
                    pass
            ws_datos.cell(row=row_num, column=2, value=fecha)
            
            ws_datos.cell(row=row_num, column=3, value=pedido.get('cliente_nombre', ''))
            ws_datos.cell(row=row_num, column=4, value=pedido.get('cliente_cuit', ''))
            ws_datos.cell(row=row_num, column=5, value=pedido.get('cliente_telefono', ''))
            ws_datos.cell(row=row_num, column=6, value=pedido.get('cliente_email', ''))
            ws_datos.cell(row=row_num, column=7, value=pedido.get('metodo_entrega', ''))
            ws_datos.cell(row=row_num, column=8, value=pedido.get('envio_direccion', ''))
            ws_datos.cell(row=row_num, column=9, value=pedido.get('envio_localidad', ''))
            ws_datos.cell(row=row_num, column=10, value=pedido.get('envio_provincia', ''))
            ws_datos.cell(row=row_num, column=11, value=pedido.get('envio_cp', ''))
            ws_datos.cell(row=row_num, column=12, value=pedido.get('envio_nombre_destinatario', ''))
            ws_datos.cell(row=row_num, column=13, value=pedido.get('envio_referencias', ''))
            
            # Total como número
            total = pedido.get('total', 0)
            ws_datos.cell(row=row_num, column=14, value=total)
            
            ws_datos.cell(row=row_num, column=15, value=pedido.get('estado', 'pendiente'))
        
        # Ajustar ancho de columnas datos
        ws_datos.column_dimensions['A'].width = 8
        ws_datos.column_dimensions['B'].width = 18
        ws_datos.column_dimensions['C'].width = 25
        ws_datos.column_dimensions['D'].width = 15
        ws_datos.column_dimensions['E'].width = 15
        ws_datos.column_dimensions['F'].width = 30
        ws_datos.column_dimensions['G'].width = 15
        ws_datos.column_dimensions['H'].width = 35
        ws_datos.column_dimensions['I'].width = 20
        ws_datos.column_dimensions['J'].width = 15
        ws_datos.column_dimensions['K'].width = 10
        ws_datos.column_dimensions['L'].width = 25
        ws_datos.column_dimensions['M'].width = 30
        ws_datos.column_dimensions['N'].width = 12
        ws_datos.column_dimensions['O'].width = 12
        
        # ===== ARCHIVO 2: PRODUCTOS DEL PEDIDO =====
        wb_productos = Workbook()
        ws_productos = wb_productos.active
        ws_productos.title = "Productos Pedidos"
        
        # Encabezados para productos
        prod_headers = ["Pedido ID", "Código", "Cantidad"]
        for col_num, header in enumerate(prod_headers, 1):
            cell = ws_productos.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Agregar productos
        prod_row = 2
        for pedido in pedidos:
            try:
                productos = json.loads(pedido.get('productos', '[]'))
                for prod in productos:
                    ws_productos.cell(row=prod_row, column=1, value=pedido.get('id', ''))
                    ws_productos.cell(row=prod_row, column=2, value=prod.get('codigo', ''))
                    ws_productos.cell(row=prod_row, column=3, value=prod.get('cantidad', 1))
                    prod_row += 1
            except:
                pass
        
        # Ajustar ancho de columnas productos
        ws_productos.column_dimensions['A'].width = 10
        ws_productos.column_dimensions['B'].width = 15
        ws_productos.column_dimensions['C'].width = 10
        
        # Guardar ambos archivos en memoria
        output_datos = BytesIO()
        wb_datos.save(output_datos)
        output_datos.seek(0)
        
        output_productos = BytesIO()
        wb_productos.save(output_productos)
        output_productos.seek(0)
        
        # Crear ZIP con ambos archivos
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
            zip_file.writestr(f'pedidos_datos_{fecha_actual}.xlsx', output_datos.getvalue())
            zip_file.writestr(f'pedidos_productos_{fecha_actual}.xlsx', output_productos.getvalue())
        
        zip_buffer.seek(0)
        
        # Generar nombre de archivo con fecha
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"pedidos_rmkits_{fecha_actual}.zip"
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error al exportar pedidos: {e}")
        flash('Error al exportar pedidos', 'error')
        return redirect(url_for('admin_pedidos'))


@app.route("/admin/pedido/<int:id>/eliminar", methods=["POST"])
@login_required
def admin_eliminar_pedido(id):
    """Eliminar un pedido individual"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pedido WHERE id = ?", (id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return jsonify({'success': True})
            else:
                return jsonify({'success': False, 'error': 'Pedido no encontrado'}), 404
    except Exception as e:
        logger.error(f"Error al eliminar pedido: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/admin/pedidos/importar", methods=["POST"])
@login_required
def admin_importar_pedidos():
    """Importar pedidos desde dos archivos Excel (datos y productos)"""
    try:
        # Validar que se recibieron ambos archivos
        if 'archivo_datos' not in request.files or 'archivo_productos' not in request.files:
            flash('Debes seleccionar ambos archivos: datos y productos', 'error')
            return redirect(url_for('admin_pedidos'))
        
        file_datos = request.files['archivo_datos']
        file_productos = request.files['archivo_productos']
        
        if file_datos.filename == '' or file_productos.filename == '':
            flash('Debes seleccionar ambos archivos', 'error')
            return redirect(url_for('admin_pedidos'))
        
        if not file_datos.filename.endswith(('.xlsx', '.xls')) or not file_productos.filename.endswith(('.xlsx', '.xls')):
            flash('Ambos archivos deben ser Excel (.xlsx o .xls)', 'error')
            return redirect(url_for('admin_pedidos'))
        
        # Leer archivo de datos
        wb_datos = load_workbook(file_datos)
        ws_datos = wb_datos.active
        
        # Leer archivo de productos
        wb_productos = load_workbook(file_productos)
        ws_productos = wb_productos.active
        
        # Primero, construir un diccionario de productos por pedido
        productos_por_pedido = {}
        for row in ws_productos.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Si no hay Pedido ID, saltar
                continue
            
            pedido_id = int(row[0])
            codigo = row[1] if row[1] else ''
            cantidad = int(row[2]) if row[2] else 1
            
            if pedido_id not in productos_por_pedido:
                productos_por_pedido[pedido_id] = []
            
            if codigo:
                productos_por_pedido[pedido_id].append({'codigo': codigo, 'cantidad': cantidad})
        
        # Ahora importar los datos de los pedidos
        pedidos_importados = 0
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Leer desde la fila 2 (la 1 son encabezados)
            for row in ws_datos.iter_rows(min_row=2, values_only=True):
                # Verificar que la fila tenga datos
                if not row[0]:  # Si no hay ID, saltar
                    continue
                
                # Extraer datos de la fila
                # [ID, Fecha, Cliente, CUIT, Teléfono, Email, Método Entrega, Dirección Envío, 
                #  Localidad, Provincia, CP, Destinatario, Referencias, Total, Estado]
                
                pedido_id = int(row[0]) if row[0] else None
                
                fecha_str = row[1] if row[1] else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(fecha_str, datetime):
                    fecha_str = fecha_str.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(fecha_str, str):
                    # Intentar parsear si es string en formato dd/mm/yyyy
                    try:
                        fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y %H:%M')
                        fecha_str = fecha_obj.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        try:
                            fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y')
                            fecha_str = fecha_obj.strftime('%Y-%m-%d %H:%M:%S')
                        except:
                            fecha_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Obtener productos para este pedido y buscar sus datos en la BD
                productos_json = '[]'
                if pedido_id in productos_por_pedido:
                    productos_data = productos_por_pedido[pedido_id]
                    productos = []
                    
                    for prod_data in productos_data:
                        codigo = prod_data['codigo']
                        cantidad = prod_data['cantidad']
                        # Buscar el producto en la base de datos
                        cursor.execute("SELECT codigo, titulo, precio FROM producto WHERE codigo = ?", (codigo,))
                        prod_row = cursor.fetchone()
                        if prod_row:
                            productos.append({
                                'codigo': prod_row[0],
                                'titulo': prod_row[1],
                                'cantidad': cantidad,  # Cantidad del Excel
                                'precio': prod_row[2]
                            })
                    
                    productos_json = json.dumps(productos)
                
                # Insertar o reemplazar el pedido respetando el ID del Excel
                # Si el ID ya existe, se actualizará; si no existe, se insertará
                cursor.execute("""
                    INSERT OR REPLACE INTO pedido (id, fecha, cliente_nombre, cliente_cuit, cliente_telefono, 
                                       cliente_email, metodo_entrega, envio_direccion, envio_localidad,
                                       envio_provincia, envio_cp, envio_nombre_destinatario, 
                                       envio_referencias, productos, total, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    pedido_id,  # ID del Excel (número de pedido)
                    fecha_str,
                    row[2] if row[2] else '',  # cliente_nombre
                    row[3] if row[3] else '',  # cliente_cuit
                    row[4] if row[4] else '',  # cliente_telefono
                    row[5] if row[5] else '',  # cliente_email
                    row[6] if row[6] else 'retiro',  # metodo_entrega
                    row[7] if row[7] else '',  # envio_direccion
                    row[8] if row[8] else '',  # envio_localidad
                    row[9] if row[9] else '',  # envio_provincia
                    row[10] if row[10] else '',  # envio_cp
                    row[11] if row[11] else '',  # envio_nombre_destinatario
                    row[12] if row[12] else '',  # envio_referencias
                    productos_json,  # productos desde el segundo Excel
                    row[13] if row[13] else 0,  # total
                    row[14] if row[14] else 'pendiente'  # estado
                ))
                pedidos_importados += 1
            
            conn.commit()
        
        flash(f'✅ {pedidos_importados} pedido(s) importado(s) correctamente', 'success')
        return redirect(url_for('admin_pedidos'))
        
    except Exception as e:
        logger.error(f"Error al importar pedidos: {e}")
        flash(f'Error al importar pedidos: {str(e)}', 'error')
        return redirect(url_for('admin_pedidos'))


@app.route("/admin/pedido/<int:id>/estado", methods=["POST"])
@login_required
def admin_pedido_estado(id):
    """Cambiar estado de un pedido"""
    try:
        data = request.get_json()
        nuevo_estado = data.get('estado')
        
        estados_validos = ['pendiente', 'recibido', 'confirmado', 'preparando', 'pagado', 'completado', 'cancelado', 'impreso', 'señado', 'preparado']
        if nuevo_estado not in estados_validos:
            return jsonify({'success': False, 'error': 'Estado inválido'}), 400
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE pedido SET estado = ? WHERE id = ?", (nuevo_estado, id))
            conn.commit()
            
            if cursor.rowcount == 0:
                return jsonify({'success': False, 'error': 'Pedido no encontrado'}), 404
        
        return jsonify({'success': True})
    
    except Exception as e:
        logger.error(f"Error al cambiar estado del pedido: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/admin/pedidos/limpiar", methods=["POST"])
@login_required
def admin_limpiar_pedidos():
    """Eliminar todos los pedidos y resetear el autoincremento"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pedido")
            cantidad = cursor.rowcount
            # Resetear el autoincremento de SQLite
            cursor.execute("DELETE FROM sqlite_sequence WHERE name='pedido'")
            conn.commit()
        
        flash(f'{cantidad} pedido(s) eliminado(s) y base de datos reseteada', 'success')
        return jsonify({'success': True, 'cantidad': cantidad})
    
    except Exception as e:
        logger.error(f"Error al limpiar pedidos: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/guardar-pedido", methods=["POST"])
def guardar_pedido():
    """Guardar pedido en la base de datos"""
    try:
        data = request.get_json()

        metodo_entrega = (data.get('metodo_entrega') or '').strip()
        if metodo_entrega == 'envio':
            dni_destinatario = (data.get('envio_dni_destinatario') or data.get('envio_cuit_destinatario') or '').strip()
            if not dni_destinatario:
                return jsonify({'success': False, 'error': 'El DNI del destinatario es obligatorio para envíos'}), 400
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Verificar si la tabla existe y tiene todas las columnas necesarias
            try:
                cursor.execute("SELECT envio_direccion FROM pedido LIMIT 1")
            except sqlite3.OperationalError:
                # La tabla no existe o no tiene las columnas nuevas, recrearla
                logger.info("Recreando tabla de pedidos con nuevas columnas")
                cursor.execute("DROP TABLE IF EXISTS pedido")
                cursor.execute("""
                    CREATE TABLE pedido (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        cliente_nombre TEXT NOT NULL,
                        cliente_cuit TEXT,
                        cliente_telefono TEXT,
                        cliente_email TEXT,
                        cliente_direccion TEXT,
                        metodo_entrega TEXT,
                        envio_direccion TEXT,
                        envio_localidad TEXT,
                        envio_provincia TEXT,
                        envio_cp TEXT,
                        envio_nombre_destinatario TEXT,
                        envio_dni_destinatario TEXT,
                        envio_referencias TEXT,
                        productos TEXT NOT NULL,
                        total REAL NOT NULL,
                        estado TEXT DEFAULT 'pendiente'
                    )
                """)
                # Configurar el AUTOINCREMENT para que empiece en 1200
                cursor.execute("UPDATE sqlite_sequence SET seq = 1199 WHERE name = 'pedido'")
                # Si no existe entrada en sqlite_sequence, insertarla
                cursor.execute("INSERT OR IGNORE INTO sqlite_sequence (name, seq) VALUES ('pedido', 1199)")

            try:
                cursor.execute("SELECT envio_dni_destinatario FROM pedido LIMIT 1")
            except sqlite3.OperationalError:
                cursor.execute("ALTER TABLE pedido ADD COLUMN envio_dni_destinatario TEXT")
            
            # Insertar pedido
            cursor.execute("""
                INSERT INTO pedido (cliente_nombre, cliente_cuit, cliente_telefono, cliente_email, 
                                   cliente_direccion, metodo_entrega, envio_direccion, envio_localidad,
                                   envio_provincia, envio_cp, envio_nombre_destinatario, envio_dni_destinatario,
                                   envio_referencias, productos, total)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('nombre'),
                data.get('cuit'),
                data.get('telefono'),
                data.get('email'),
                data.get('direccion'),
                data.get('metodo_entrega'),
                data.get('envio_direccion'),
                data.get('envio_localidad'),
                data.get('envio_provincia'),
                data.get('envio_cp'),
                data.get('envio_nombre_destinatario'),
                data.get('envio_dni_destinatario') or data.get('envio_cuit_destinatario'),
                data.get('envio_referencias'),
                data.get('productos'),  # JSON string con los productos
                data.get('total')
            ))
            
            pedido_id = cursor.lastrowid
            conn.commit()
        
        # Enviar email de confirmación al cliente
        if data.get('email'):
            datos_envio = None
            if data.get('metodo_entrega') == 'envio':
                datos_envio = {
                    'direccion': data.get('envio_direccion'),
                    'localidad': data.get('envio_localidad'),
                    'provincia': data.get('envio_provincia'),
                    'cp': data.get('envio_cp'),
                    'nombre_destinatario': data.get('envio_nombre_destinatario'),
                    'dni_destinatario': data.get('envio_dni_destinatario') or data.get('envio_cuit_destinatario'),
                    'referencias': data.get('envio_referencias')
                }
            
            enviar_email_confirmacion(
                pedido_id=pedido_id,
                cliente_nombre=data.get('nombre'),
                cliente_email=data.get('email'),
                productos_json=data.get('productos'),
                total=data.get('total'),
                metodo_entrega=data.get('metodo_entrega', 'retiro'),
                datos_envio=datos_envio
            )
        
        return jsonify({'success': True, 'pedido_id': pedido_id})
    
    except Exception as e:
        logger.error(f"Error al guardar pedido: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/gracias")
def gracias():
    """Página de agradecimiento después de completar un pedido"""
    pedido_id = request.args.get('pedido_id')
    return render_template('gracias.html', pedido_id=pedido_id)


@app.route("/admin/lista-precios")
@login_required
def admin_lista_precios():
    """Lista de precios para imprimir"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM producto WHERE stock > 0 ORDER BY codigo DESC")
            productos = [dict(row) for row in cursor.fetchall()]
            
            # Obtener categorías únicas
            cursor.execute("SELECT DISTINCT categoria FROM producto WHERE categoria IS NOT NULL AND categoria != '' ORDER BY categoria")
            categorias = [row['categoria'] for row in cursor.fetchall()]
            
            return render_template("admin/lista_precios.html", productos=productos, categorias=categorias, now=datetime.now())
    except Exception as e:
        logger.error(f"Error al obtener lista de precios: {e}")
        flash('Error al cargar lista de precios', 'error')
        return redirect(url_for('admin_productos'))



@app.route("/admin/producto/nuevo", methods=["GET", "POST"])
@login_required
def admin_producto_nuevo():
    """Crear nuevo producto"""
    if request.method == "POST":
        try:
            # Generar código automáticamente primero
            codigo_automatico = generar_codigo_producto()
            
            # Manejar imagen
            imagen_filename = ""
            if 'imagen' in request.files:
                file = request.files['imagen']
                if file and file.filename and allowed_file(file.filename):
                    # Obtener la extensión del archivo original
                    extension = os.path.splitext(file.filename)[1].lower()
                    # Renombrar con el código del producto
                    imagen_filename = f"{codigo_automatico}{extension}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], imagen_filename)
                    file.save(filepath)
            
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO producto (codigo, titulo, descripcion, precio, minimo, multiplo, stock, imagen, categoria, activo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    codigo_automatico,  # Usar código generado automáticamente
                    request.form.get('titulo'),
                    request.form.get('descripcion', ''),
                    float(request.form.get('precio', 0)),
                    int(request.form.get('minimo', 1)),
                    int(request.form.get('multiplo', 1)),
                    int(request.form.get('stock', 0)),
                    imagen_filename,
                    request.form.get('categoria', ''),
                    1  # Por defecto activo
                ))
                
                # Obtener el ID del producto recién insertado
                producto_id = cursor.lastrowid
                
                # Agregar a la tabla de productos nuevos
                cursor.execute("""
                    INSERT INTO producto_nuevo (producto_id)
                    VALUES (?)
                """, (producto_id,))
                
                conn.commit()
            
            flash(f'Producto creado exitosamente con código {codigo_automatico}', 'success')
            return redirect(url_for('admin_productos'))
        except Exception as e:
            logger.error(f"Error al crear producto: {e}")
            flash(f'Error al crear producto: {str(e)}', 'error')
    
    # Generar el código que se usará para el nuevo producto
    nuevo_codigo = generar_codigo_producto()
    
    # Obtener categorías de la base de datos
    categorias = []
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM categoria ORDER BY nombre")
            categorias = [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        logger.error(f"Error al obtener categorías: {e}")
    
    return render_template("admin/producto_form.html", producto=None, action="Crear", nuevo_codigo=nuevo_codigo, categorias=categorias)


@app.route("/admin/producto/<int:id>/editar", methods=["GET", "POST"])
@login_required
def admin_producto_editar(id):
    """Editar producto existente"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            if request.method == "POST":
                # Obtener el código del producto
                codigo_producto = request.form.get('codigo')
                
                # Manejar imagen
                imagen_filename = request.form.get('imagen_actual', '')
                if 'imagen' in request.files:
                    file = request.files['imagen']
                    if file and file.filename and allowed_file(file.filename):
                        # Obtener la extensión del archivo original
                        extension = os.path.splitext(file.filename)[1].lower()
                        # Renombrar con el código del producto
                        imagen_filename = f"{codigo_producto}{extension}"
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], imagen_filename)
                        file.save(filepath)
                        
                        # Eliminar imagen anterior si existe y es diferente
                        imagen_anterior = request.form.get('imagen_actual', '')
                        if imagen_anterior and imagen_anterior != imagen_filename:
                            ruta_anterior = os.path.join(app.config['UPLOAD_FOLDER'], imagen_anterior)
                            if os.path.exists(ruta_anterior):
                                try:
                                    os.remove(ruta_anterior)
                                except Exception as e:
                                    logger.warning(f"No se pudo eliminar la imagen anterior: {e}")
                
                cursor.execute("""
                    UPDATE producto 
                    SET codigo=?, titulo=?, descripcion=?, precio=?, minimo=?, multiplo=?, stock=?, imagen=?, categoria=?
                    WHERE id=?
                """, (
                    request.form.get('codigo'),
                    request.form.get('titulo'),
                    request.form.get('descripcion', ''),
                    float(request.form.get('precio', 0)),
                    int(request.form.get('minimo', 1)),
                    int(request.form.get('multiplo', 1)),
                    int(request.form.get('stock', 0)),
                    imagen_filename,
                    request.form.get('categoria', ''),
                    id
                ))
                conn.commit()
                
                flash('Producto actualizado exitosamente', 'success')
                return redirect(url_for('admin_productos'))
            
            # GET - mostrar formulario
            cursor.execute("SELECT * FROM producto WHERE id=?", (id,))
            producto = dict(cursor.fetchone())
            
            # Obtener categorías de la base de datos
            cursor.execute("SELECT * FROM categoria ORDER BY nombre")
            categorias = [dict(row) for row in cursor.fetchall()]
            
            return render_template("admin/producto_form.html", producto=producto, action="Editar", categorias=categorias)
    
    except Exception as e:
        logger.error(f"Error al editar producto: {e}")
        flash(f'Error al editar producto: {str(e)}', 'error')
        return redirect(url_for('admin_productos'))


@app.route("/admin/producto/actualizar-precio", methods=["POST"])
@login_required
def admin_producto_actualizar_precio():
    """Actualizar precio de un producto vía AJAX"""
    try:
        data = request.get_json()
        producto_id = data.get('producto_id')
        nuevo_precio = data.get('precio')
        
        if not producto_id or nuevo_precio is None:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        if nuevo_precio < 0:
            return jsonify({'success': False, 'error': 'El precio no puede ser negativo'}), 400
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE producto SET precio = ? WHERE id = ?", (nuevo_precio, producto_id))
            conn.commit()
            
            if cursor.rowcount == 0:
                return jsonify({'success': False, 'error': 'Producto no encontrado'}), 404
        
        return jsonify({'success': True, 'precio': nuevo_precio})
    
    except Exception as e:
        logger.error(f"Error al actualizar precio: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/admin/producto/toggle-activo", methods=["POST"])
@login_required
def admin_producto_toggle_activo():
    """Activar/desactivar un producto vía AJAX"""
    try:
        data = request.get_json()
        producto_id = data.get('producto_id')
        activo = data.get('activo')
        
        if not producto_id or activo is None:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE producto SET activo = ? WHERE id = ?", (activo, producto_id))
            conn.commit()
            
            if cursor.rowcount == 0:
                return jsonify({'success': False, 'error': 'Producto no encontrado'}), 404
        
        return jsonify({'success': True, 'activo': activo})
    
    except Exception as e:
        logger.error(f"Error al actualizar estado del producto: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/admin/producto/<int:id>/eliminar", methods=["POST"])
@login_required
def admin_producto_eliminar(id):
    """Eliminar producto"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM producto WHERE id=?", (id,))
            conn.commit()
        
        flash('Producto eliminado exitosamente', 'success')
    except Exception as e:
        logger.error(f"Error al eliminar producto: {e}")
        flash(f'Error al eliminar producto: {str(e)}', 'error')
    
    return redirect(url_for('admin_productos'))


@app.route("/admin/api/productos")
@login_required
def admin_api_productos():
    """API para obtener productos (para tablas dinámicas)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM producto ORDER BY codigo DESC")
            productos = [dict(row) for row in cursor.fetchall()]
            return jsonify(productos)
    except Exception as e:
        logger.error(f"Error en API productos: {e}")
        return jsonify({"error": str(e)}), 500


# =============================================================================
# GESTIÓN DE CATEGORÍAS
# =============================================================================

@app.route("/admin/categorias")
@login_required
def admin_categorias():
    """Muestra la página de gestión de categorías"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM categoria ORDER BY nombre")
            categorias = [dict(row) for row in cursor.fetchall()]
            
            # Contar productos por categoría
            for categoria in categorias:
                cursor.execute("SELECT COUNT(*) as count FROM producto WHERE categoria = ?", (categoria['nombre'],))
                categoria['num_productos'] = cursor.fetchone()['count']
            
            return render_template("admin/categorias.html", categorias=categorias)
    except Exception as e:
        logger.error(f"Error al obtener categorías: {e}")
        flash('Error al cargar categorías', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route("/admin/categoria/nueva", methods=["POST"])
@login_required
def admin_categoria_nueva():
    """Crear nueva categoría"""
    try:
        nombre = request.form.get('nombre', '').strip()
        
        if not nombre:
            flash('El nombre de la categoría es requerido', 'error')
            return redirect(url_for('admin_categorias'))
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO categoria (nombre) VALUES (?)", (nombre,))
            conn.commit()
        
        flash('Categoría creada exitosamente', 'success')
    except sqlite3.IntegrityError:
        flash('Ya existe una categoría con ese nombre', 'error')
    except Exception as e:
        logger.error(f"Error al crear categoría: {e}")
        flash(f'Error al crear categoría: {str(e)}', 'error')
    
    return redirect(url_for('admin_categorias'))


@app.route("/admin/categoria/<int:id>/editar", methods=["POST"])
@login_required
def admin_categoria_editar(id):
    """Editar categoría existente"""
    try:
        nuevo_nombre = request.form.get('nombre', '').strip()
        
        if not nuevo_nombre:
            flash('El nombre de la categoría es requerido', 'error')
            return redirect(url_for('admin_categorias'))
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Obtener el nombre anterior
            cursor.execute("SELECT nombre FROM categoria WHERE id = ?", (id,))
            resultado = cursor.fetchone()
            if not resultado:
                flash('Categoría no encontrada', 'error')
                return redirect(url_for('admin_categorias'))
            
            nombre_anterior = resultado['nombre']
            
            # Actualizar el nombre de la categoría
            cursor.execute("UPDATE categoria SET nombre = ? WHERE id = ?", (nuevo_nombre, id))
            
            # Actualizar todos los productos que tenían la categoría anterior
            cursor.execute("UPDATE producto SET categoria = ? WHERE categoria = ?", (nuevo_nombre, nombre_anterior))
            
            conn.commit()
        
        flash('Categoría actualizada exitosamente', 'success')
    except sqlite3.IntegrityError:
        flash('Ya existe una categoría con ese nombre', 'error')
    except Exception as e:
        logger.error(f"Error al editar categoría: {e}")
        flash(f'Error al editar categoría: {str(e)}', 'error')
    
    return redirect(url_for('admin_categorias'))


@app.route("/admin/categoria/<int:id>/eliminar", methods=["POST"])
@login_required
def admin_categoria_eliminar(id):
    """Eliminar categoría (con reasignación de productos)"""
    try:
        categoria_destino = request.form.get('categoria_destino', '').strip()
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Obtener el nombre de la categoría a eliminar
            cursor.execute("SELECT nombre FROM categoria WHERE id = ?", (id,))
            resultado = cursor.fetchone()
            if not resultado:
                flash('Categoría no encontrada', 'error')
                return redirect(url_for('admin_categorias'))
            
            nombre_categoria = resultado['nombre']
            
            # Verificar si hay productos con esta categoría
            cursor.execute("SELECT COUNT(*) as count FROM producto WHERE categoria = ?", (nombre_categoria,))
            num_productos = cursor.fetchone()['count']
            
            if num_productos > 0:
                if not categoria_destino:
                    flash(f'Esta categoría tiene {num_productos} producto(s). Debes seleccionar una categoría de destino', 'error')
                    return redirect(url_for('admin_categorias'))
                
                # Reasignar productos a la categoría destino
                cursor.execute("UPDATE producto SET categoria = ? WHERE categoria = ?", (categoria_destino, nombre_categoria))
            
            # Eliminar la categoría
            cursor.execute("DELETE FROM categoria WHERE id = ?", (id,))
            conn.commit()
        
        flash('Categoría eliminada exitosamente', 'success')
    except Exception as e:
        logger.error(f"Error al eliminar categoría: {e}")
        flash(f'Error al eliminar categoría: {str(e)}', 'error')
    
    return redirect(url_for('admin_categorias'))


@app.route("/admin/productos-nuevos")
@login_required
def admin_productos_nuevos():
    """Muestra la tabla de productos nuevos"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.*, pn.fecha_agregado 
                FROM producto p
                INNER JOIN producto_nuevo pn ON p.id = pn.producto_id
                ORDER BY pn.fecha_agregado DESC
            """)
            productos_nuevos = [dict(row) for row in cursor.fetchall()]
            
            return render_template("admin/productos_nuevos.html", productos=productos_nuevos)
    except Exception as e:
        logger.error(f"Error al obtener productos nuevos: {e}")
        flash('Error al cargar productos nuevos', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route("/admin/productos-nuevos/descargar-imagenes")
@login_required
def admin_descargar_imagenes_nuevos():
    """Descarga un ZIP con todas las imágenes de productos nuevos"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.codigo, p.imagen 
                FROM producto p
                INNER JOIN producto_nuevo pn ON p.id = pn.producto_id
                WHERE p.imagen IS NOT NULL AND p.imagen != ''
            """)
            productos = cursor.fetchall()
        
        if not productos:
            flash('No hay productos nuevos con imágenes', 'warning')
            return redirect(url_for('admin_productos_nuevos'))
        
        # Crear un archivo ZIP en memoria
        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for producto in productos:
                codigo = producto['codigo']
                imagen = producto['imagen']
                imagen_path = os.path.join(app.config['UPLOAD_FOLDER'], imagen)
                
                if os.path.exists(imagen_path):
                    # Obtener la extensión del archivo
                    extension = os.path.splitext(imagen)[1].lower()
                    # Agregar archivo al ZIP con el código del producto como nombre
                    nombre_archivo = f"{codigo}{extension}"
                    zipf.write(imagen_path, nombre_archivo)
        
        memory_file.seek(0)
        
        # Generar nombre del archivo con fecha
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_zip = f"imagenes_productos_nuevos_{fecha_actual}.zip"
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=nombre_zip
        )
    
    except Exception as e:
        logger.error(f"Error al descargar imágenes: {e}")
        flash(f'Error al descargar imágenes: {str(e)}', 'error')
        return redirect(url_for('admin_productos_nuevos'))


@app.route("/admin/productos-nuevos/<int:id>/quitar", methods=["POST"])
@login_required
def admin_quitar_producto_nuevo(id):
    """Quita un producto de la tabla de productos nuevos"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM producto_nuevo WHERE producto_id = ?", (id,))
            conn.commit()
        
        flash('Producto quitado de la lista de nuevos', 'success')
    except Exception as e:
        logger.error(f"Error al quitar producto nuevo: {e}")
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('admin_productos_nuevos'))


# Ruta para servir imágenes desde almacenamiento persistente
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Sirve imágenes desde la carpeta de almacenamiento persistente"""
    return send_from_directory(Config.UPLOAD_FOLDER, filename)


if __name__ == "__main__":
    app.run(debug=Config.DEBUG)
